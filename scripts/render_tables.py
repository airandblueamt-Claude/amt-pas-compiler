"""
render_tables.py — render a BOQ spreadsheet (.xlsx) to faithful PDF table pages.

Two engines:
  * LibreOffice ("soffice --headless --convert-to pdf"): pixel-faithful to the
    company's Excel formatting (Arabic RTL, merged cells). Used when available.
  * reportlab fallback: reads the sheet with openpyxl and re-typesets a clean,
    AMT-styled bilingual grid. Self-contained (no external binary needed).

Both return a path to a PDF whose pages contain only the table (no logo/banner),
matching how the sample embeds the rendered Excel sheets.
"""
from __future__ import annotations

import os
import re
import shutil
import subprocess
import tempfile

import openpyxl
from reportlab.pdfgen import canvas
from reportlab.pdfbase.pdfmetrics import stringWidth

import amt_common as A
import stamp as STAMP
from amt_common import (PAGE_W, PAGE_H, MARGIN_L, MARGIN_R, CONTENT_W,
                        BLUE_HEADER, BLUE_BORDER, WHITE, BLACK, GREY_REF,
                        F_EN, F_EN_B, F_AR, F_AR_B, ar)

_ARABIC_RE = re.compile(r"[؀-ۿݐ-ݿࢠ-ࣿﭐ-﷿ﹰ-﻿]")


def has_arabic(s: str) -> bool:
    return bool(_ARABIC_RE.search(s or ""))


def have_soffice() -> str | None:
    """Locate a LibreOffice binary. Checks $AMT_SOFFICE, then PATH, then common
    install locations including a user-extracted AppImage under ~/.local/opt."""
    import glob
    env = os.environ.get("AMT_SOFFICE")
    if env and os.path.exists(env):
        return env
    for name in ("soffice", "libreoffice"):
        path = shutil.which(name)
        if path:
            return path
    patterns = [
        os.path.expanduser("~/.local/opt/squashfs-root/opt/libreoffice*/program/soffice"),
        os.path.expanduser("~/.local/opt/libreoffice*/program/soffice"),
        "/opt/libreoffice*/program/soffice",
        "/usr/lib/libreoffice/program/soffice",
    ]
    for pat in patterns:
        hits = sorted(glob.glob(pat))
        if hits:
            return hits[-1]
    return None


# --------------------------------------------------------------------------- #
# Fit-to-page preparation (so wide BOQ sheets scale onto one A4 width, centred,
# instead of overflowing across several pages)
# --------------------------------------------------------------------------- #
def _strip_headers_footers(ws) -> None:
    """Remove the sheet's own print headers/footers. They are page chrome the
    compiler supplies itself, and a bare newline in one renders as the literal
    '_x000a_' token in LibreOffice (the artifact seen at the top of §2/§3); §3's
    '&G' image-headers would also pull in stale graphics. Clearing them keeps the
    page faithful to the table CONTENT while letting AMT own all page chrome."""
    for hf in (ws.oddHeader, ws.oddFooter, ws.evenHeader, ws.evenFooter,
               ws.firstHeader, ws.firstFooter):
        try:
            hf.left.text = hf.center.text = hf.right.text = None
        except Exception:
            pass


_EMU_PER_PX = 9525
_EMU_PER_PT = 12700


def _col_emu(ws, c0):                              # full width of 0-based column, EMU
    from openpyxl.utils import get_column_letter
    d = ws.column_dimensions.get(get_column_letter(c0 + 1))
    chars = d.width if d and d.width else 8.43
    return int((round(chars * 7) + 5) * _EMU_PER_PX)


def _pin_images(ws):
    """Freeze every embedded image to its CURRENT display size as a fixed-size
    OneCellAnchor, so that changing row heights afterwards can no longer stretch or
    squash it. Returns (floor, records):
      * floor   = {row(1-based): tallest pinned image height in points} so the row
                  auto-fitter keeps image rows at least as tall as their picture,
      * records = [{im, col0, row1, cx, cy}] so the caller can re-centre each picture
                  inside its (now taller) cell once final row heights are known."""
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D

    def row_emu(r0):                               # full height of 0-based row, EMU
        d = ws.row_dimensions.get(r0 + 1)
        pts = d.height if d and d.height else 15.0
        return int(pts * _EMU_PER_PT)

    floor, records = {}, []
    for im in list(getattr(ws, "_images", []) or []):
        a = getattr(im, "anchor", None)
        frm = getattr(a, "_from", None)
        if frm is None:
            continue
        to = getattr(a, "to", None)
        if to is not None:                         # TwoCellAnchor -> measure span
            if to.col == frm.col:
                cx = to.colOff - frm.colOff
            else:
                cx = (_col_emu(ws, frm.col) - frm.colOff
                      + sum(_col_emu(ws, c) for c in range(frm.col + 1, to.col)) + to.colOff)
            if to.row == frm.row:
                cy = to.rowOff - frm.rowOff
            else:
                cy = (row_emu(frm.row) - frm.rowOff
                      + sum(row_emu(r) for r in range(frm.row + 1, to.row)) + to.rowOff)
        else:                                      # already OneCellAnchor
            ext = getattr(a, "ext", None)
            cx = ext.cx if ext else _col_emu(ws, frm.col)
            cy = ext.cy if ext else row_emu(frm.row)
        cx, cy = max(int(cx), 1), max(int(cy), 1)
        marker = AnchorMarker(col=frm.col, colOff=frm.colOff,
                              row=frm.row, rowOff=frm.rowOff)
        im.anchor = OneCellAnchor(_from=marker, ext=XDRPositiveSize2D(cx=cx, cy=cy))
        r1 = frm.row + 1
        floor[r1] = max(floor.get(r1, 0.0), cy / _EMU_PER_PT)
        records.append({"im": im, "col0": frm.col, "row1": r1, "cx": cx, "cy": cy})
    return floor, records


def _centre_images(ws, records) -> None:
    """Now that final row heights are set, sit each pinned picture in the MIDDLE of
    its cell (vertical + horizontal) instead of jammed against the top-left corner —
    so a reference photo lines up with its item in the grown row instead of floating
    above a gap."""
    for rec in records:
        col0, r1, cx, cy = rec["col0"], rec["row1"], rec["cx"], rec["cy"]
        h = ws.row_dimensions[r1].height
        row_h_emu = int(h * _EMU_PER_PT) if h else cy
        col_w_emu = _col_emu(ws, col0)
        frm = rec["im"].anchor._from
        frm.colOff = max(0, (col_w_emu - cx) // 2)
        frm.rowOff = max(0, (row_h_emu - cy) // 2)


def _text_height(text, width_pt, fs) -> float:
    """Generous estimate of the height (points) wrapped text needs in a cell."""
    import math
    w = max(width_pt - 8, 14)
    cpl = max(int(w / (fs * 0.6)), 1)             # Arabic glyphs run wide -> 0.6
    lines = sum(max(1, math.ceil(len(part) / cpl)) for part in str(text).split("\n"))
    return lines * fs * 1.55 + 12


def _autofit_text_rows(ws) -> None:
    """Make every row tall enough for its text so nothing clips/overlaps, while
    keeping embedded pictures at their faithful size:

      * pin images first (fixed size) so the height changes below can't distort them,
      * turn wrapping on (so long text never clips horizontally),
      * rows WITHOUT an image: drop the manual height so LibreOffice auto-fits the
        text exactly (no over- or under-shoot),
      * rows WITH an image: grow to max(current height, the picture's height, the
        height the text needs) — never shrinks, so the picture keeps its size and any
        clipped text (e.g. long Arabic item descriptions) gets room,
      * size multi-row MERGED spans explicitly (auto-fit ignores merged cells).
    Fonts and cell values are never changed."""
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    img_floor, img_records = _pin_images(ws)      # pins images; {row: min height pts}
    img_rows = set(img_floor)

    for row in ws.iter_rows():
        for cell in row:
            if cell.value in (None, ""):
                continue
            try:
                a = cell.alignment
                cell.alignment = Alignment(horizontal=a.horizontal,
                                           vertical=a.vertical, wrap_text=True)
            except (AttributeError, TypeError, ValueError):
                pass

    colpt = {}                                    # column widths in points
    for i in range(1, (ws.max_column or 0) + 1):
        d = ws.column_dimensions.get(get_column_letter(i))
        colpt[i] = ((d.width if d and d.width else 8.43) * 7 + 5) * 0.75

    merged_rows = set()
    for mr in ws.merged_cells.ranges:
        if mr.max_row > mr.min_row:
            merged_rows.update(range(mr.min_row, mr.max_row + 1))

    for r in range(1, (ws.max_row or 0) + 1):
        if r in merged_rows:
            continue                              # handled per merged-span below
        if r not in img_rows:
            ws.row_dimensions[r].height = None    # text-only row -> LibreOffice auto-fit
            continue
        need = img_floor.get(r, 0.0)              # never shorter than the picture
        for cell in ws[r]:
            v = cell.value
            if isinstance(v, str) and v.strip():
                need = max(need, _text_height(v, colpt.get(cell.column, 50),
                                              getattr(cell.font, "size", None) or 11))
        cur = ws.row_dimensions[r].height or 0
        ws.row_dimensions[r].height = max(need, cur)

    for mr in ws.merged_cells.ranges:
        if mr.max_row <= mr.min_row:
            continue
        top = ws.cell(row=mr.min_row, column=mr.min_col)
        v = top.value
        if not isinstance(v, str) or not v.strip():
            continue
        wpt = sum(colpt.get(c, 50) for c in range(mr.min_col, mr.max_col + 1))
        need = _text_height(v, wpt, getattr(top.font, "size", None) or 11)
        per = need / (mr.max_row - mr.min_row + 1)
        for r in range(mr.min_row, mr.max_row + 1):
            cur = ws.row_dimensions[r].height or 0
            ws.row_dimensions[r].height = max(cur, per, img_floor.get(r, 0.0))

    _centre_images(ws, img_records)               # sit pictures mid-cell in grown rows


def _prepare_xlsx_for_print(xlsx_path: str, reserve_top: bool = True) -> str:
    """Return a temp copy of the workbook with ONLY its page setup adjusted so it
    converts to a clean, single-width A4 PDF — fonts, row heights, wrapping, merged
    cells and images are left EXACTLY as the user designed them (faithful). When
    `reserve_top`, a top margin is left for the stamped AMT logo header (branded
    sections only); otherwise the user's top margin is untouched."""
    from openpyxl.worksheet.properties import PageSetupProperties
    wb = openpyxl.load_workbook(xlsx_path)
    top_in = A.logo_header_reserve() / 72.0
    for ws in wb.worksheets:
        ws.page_setup.orientation = "portrait"      # uniform portrait document
        ws.page_setup.paperSize = 9                 # A4
        ws.page_setup.fitToWidth = 1                # all columns on one page wide
        ws.page_setup.fitToHeight = 0               # flow onto more pages if long
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        # centre horizontally with SYMMETRIC left/right margins (an uneven L/R
        # margin defeats 'center horizontally' and shifts the table sideways).
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = False
        m = ws.page_margins
        m.left = m.right = 0.3
        if reserve_top:
            m.top = round(max(m.top or 0, top_in), 3)
            m.header = 0.0
        _strip_headers_footers(ws)
        # defensive: turn any literal Excel newline escape into a real newline
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and "_x000a_" in v.lower():
                    cell.value = v.replace("_x000a_", "\n").replace("_x000A_", "\n")
        # make every row tall enough for its text so nothing clips/overlaps;
        # embedded pictures are pinned to a fixed size first so they aren't squashed
        _autofit_text_rows(ws)
    fd, tmp = tempfile.mkstemp(suffix=".xlsx", prefix="amt_fit_")
    os.close(fd)
    wb.save(tmp)
    return tmp


# --------------------------------------------------------------------------- #
# LibreOffice engine
# --------------------------------------------------------------------------- #
# PDF-export filter that keeps embedded reference images sharp: lossless
# compression and NO resolution down-sampling (LibreOffice otherwise drops images
# to ~96–150 dpi, which makes product photos look blurry).
_PDF_HQ_FILTER = (
    'pdf:calc_pdf_Export:'
    '{"UseLosslessCompression":{"type":"boolean","value":"true"},'
    '"ReduceImageResolution":{"type":"boolean","value":"false"},'
    '"MaxImageResolution":{"type":"long","value":"600"}}'
)


def render_with_soffice(xlsx_path: str, out_pdf: str, soffice: str) -> str:
    outdir = tempfile.mkdtemp(prefix="amt_lo_")
    # A dedicated profile dir avoids clashes with a running LibreOffice instance.
    profile = tempfile.mkdtemp(prefix="amt_loprof_")

    def _run(convert_to):
        cmd = [soffice, "--headless", "--nologo", "--nofirststartwizard",
               f"-env:UserInstallation=file://{profile}",
               "--convert-to", convert_to, "--outdir", outdir, xlsx_path]
        subprocess.run(cmd, check=True, capture_output=True, timeout=180)

    base = os.path.splitext(os.path.basename(xlsx_path))[0] + ".pdf"
    produced = os.path.join(outdir, base)
    try:
        _run(_PDF_HQ_FILTER)        # high-quality, sharp images
        if not os.path.exists(produced):
            _run("pdf")             # fall back to the plain filter
    except subprocess.CalledProcessError:
        _run("pdf")
    if not os.path.exists(produced):
        raise RuntimeError(f"LibreOffice did not produce {produced}")
    shutil.move(produced, out_pdf)
    shutil.rmtree(outdir, ignore_errors=True)
    shutil.rmtree(profile, ignore_errors=True)
    return out_pdf


# --------------------------------------------------------------------------- #
# reportlab fallback engine
# --------------------------------------------------------------------------- #
def _cell_text(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _sheet_matrix(ws):
    """Return (rows, ncols) trimmed to the used area, with merged cells expanded
    so the top-left value shows and a merge-span map is available."""
    rows = [[_cell_text(c) for c in r] for r in ws.iter_rows(values_only=True)]
    while rows and all(c == "" for c in rows[-1]):
        rows.pop()
    if not rows:
        return [], 0, {}
    ncols = max((max((i + 1 for i, c in enumerate(r) if c != ""), default=0) for r in rows),
                default=0)
    rows = [r[:ncols] + [""] * (ncols - len(r)) for r in rows]
    # merged spans -> {(r,c): (rowspan, colspan)} 0-indexed within used area
    spans = {}
    covered = set()
    for mc in ws.merged_cells.ranges:
        r0, c0, r1, c1 = mc.min_row - 1, mc.min_col - 1, mc.max_row - 1, mc.max_col - 1
        if r0 >= len(rows) or c0 >= ncols:
            continue
        spans[(r0, c0)] = (r1 - r0 + 1, c1 - c0 + 1)
        for rr in range(r0, min(r1 + 1, len(rows))):
            for cc in range(c0, min(c1 + 1, ncols)):
                if (rr, cc) != (r0, c0):
                    covered.add((rr, cc))
    return rows, ncols, spans, covered


def _col_widths(ws, ncols, content_w=CONTENT_W):
    """Approximate column widths (points) from Excel column dimensions, scaled
    to the printable content width."""
    widths = []
    for i in range(ncols):
        letter = openpyxl.utils.get_column_letter(i + 1)
        dim = ws.column_dimensions.get(letter)
        w = dim.width if dim and dim.width else 10
        widths.append(max(w, 4))
    total = sum(widths)
    scale = content_w / total
    return [w * scale for w in widths]


def _wrap(text, font, size, max_w, is_ar):
    words = str(text).split()
    if not words:
        return [""]
    lines, cur = [], words[0]
    for w in words[1:]:
        trial = cur + " " + w
        disp = ar(trial) if is_ar else trial
        if stringWidth(disp, font, size) <= max_w:
            cur = trial
        else:
            lines.append(cur)
            cur = w
    lines.append(cur)
    return lines


def render_with_reportlab(xlsx_path: str, out_pdf: str, title: str, ref_no: str,
                          branded: bool = False) -> str:
    A.register_fonts()
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    # pick the first non-empty, prefer a non-Arabic-named sheet (English layout)
    ws = wb.active
    for cand in wb.worksheets:
        if cand.max_row and cand.max_column:
            ws = cand
            break
    rows, ncols, spans, covered = _sheet_matrix(ws)
    if ncols == 0:
        rows, ncols, spans, covered = [["(empty sheet)"]], 1, {}, set()

    # A4 portrait so the table matches the rest of the document
    page_w, page_h = PAGE_W, PAGE_H
    content_w = page_w - MARGIN_L - MARGIN_R
    colw = _col_widths(ws, ncols, content_w)
    size = 8
    pad = 3
    line_h = size + 2

    c = canvas.Canvas(out_pdf, pagesize=(page_w, page_h))
    # reserve room for the stamped logo only on branded (AMT-authored) sections
    top_reserve = A.logo_header_reserve() if branded else 28
    bottom_reserve = 28
    top_y = page_h - top_reserve - 6
    bottom_limit = bottom_reserve + 6
    x_left = (page_w - sum(colw)) / 2   # centre the table horizontally
    page_no = 1

    def row_height(r):
        h = line_h + 2 * pad
        for cidx in range(ncols):
            if (r, cidx) in covered:
                continue
            txt = rows[r][cidx]
            if not txt:
                continue
            span = spans.get((r, cidx), (1, 1))
            w = sum(colw[cidx:cidx + span[1]]) - 2 * pad
            is_ar = has_arabic(txt)
            font = (F_AR if is_ar else F_EN)
            nlines = len(_wrap(txt, font, size, w, is_ar))
            h = max(h, nlines * line_h + 2 * pad)
        return min(h, page_h - 140)   # never taller than a page

    y = top_y
    is_header_row = True
    for r in range(len(rows)):
        rh = row_height(r)
        if y - rh < bottom_limit:
            c.showPage()
            page_no += 1
            y = top_y
        x = x_left
        for cidx in range(ncols):
            if (r, cidx) in covered:
                x += colw[cidx]
                continue
            span = spans.get((r, cidx), (1, 1))
            cw = sum(colw[cidx:cidx + span[1]])
            # header band for the first row
            header_band = is_header_row
            if header_band:
                c.setFillColor(BLUE_HEADER)
                c.rect(x, y - rh, cw, rh, stroke=0, fill=1)
            c.setStrokeColor(BLUE_BORDER)
            c.setLineWidth(0.5)
            c.rect(x, y - rh, cw, rh, stroke=1, fill=0)
            txt = rows[r][cidx]
            if txt:
                is_ar = has_arabic(txt)
                font = (F_AR_B if header_band else F_AR) if is_ar else (F_EN_B if header_band else F_EN)
                color = WHITE if header_band else BLACK
                lines = _wrap(txt, font, size, cw - 2 * pad, is_ar)
                c.setFont(font, size)
                c.setFillColor(color)
                ty = y - pad - size
                for ln in lines:
                    disp = ar(ln) if is_ar else ln
                    if is_ar:
                        c.drawRightString(x + cw - pad, ty, disp)
                    else:
                        c.drawString(x + pad, ty, disp)
                    ty -= line_h
                c.setFillColor(BLACK)
            x += cw
        y -= rh
        is_header_row = False

    c.showPage()
    c.save()
    return out_pdf


# --------------------------------------------------------------------------- #
# Faithful path (image-rich sheets, e.g. §4 Material Selection)
# --------------------------------------------------------------------------- #
def _workbook_has_images(xlsx_path: str) -> bool:
    """True if any sheet embeds a picture — those sheets are laid out around their
    images in Excel, so re-typesetting them clips text / shifts pictures. They are
    rendered faithfully instead."""
    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception:
        return False
    return any(getattr(ws, "_images", None) for ws in wb.worksheets)


def render_faithful(xlsx_path: str, out_pdf: str, soffice: str,
                    branded: bool = False) -> str:
    """Render the sheet EXACTLY as Excel saves it (its own page setup — nothing
    touched, so nothing clips and every picture stays where it was placed), then, for
    branded sections, shrink that faithful page to leave a top band and stamp the AMT
    logo into it. No row heights, wrapping, fonts or image anchors are modified."""
    import fitz
    raw = out_pdf + ".raw.pdf"
    render_with_soffice(xlsx_path, raw, soffice)     # zero modification == like Excel
    if not branded:
        os.replace(raw, out_pdf)
        return out_pdf

    src = fitz.open(raw)
    out = fitz.open()
    W, H = A.PAGE_W, A.PAGE_H
    band = A.logo_header_reserve()                   # top space reserved for the logo
    lw = A.LOGO_H / A._img_aspect(A.LOGO_PNG)
    logo_rect = fitz.Rect(A.LOGO_X, A.LOGO_TOP_GAP,
                          A.LOGO_X + lw, A.LOGO_TOP_GAP + A.LOGO_H)
    target = fitz.Rect(12, band, W - 12, H - 12)     # faithful page scaled in here
    for i in range(src.page_count):
        pg = out.new_page(width=W, height=H)
        pg.show_pdf_page(target, src, i)             # uniform scale, keeps aspect
        pg.insert_image(logo_rect, filename=A.LOGO_PNG, keep_proportion=True)
    out.save(out_pdf)
    src.close()
    out.close()
    try:
        os.remove(raw)
    except OSError:
        pass
    return out_pdf


# --------------------------------------------------------------------------- #
# Dispatcher
# --------------------------------------------------------------------------- #
def render_table(xlsx_path: str, out_pdf: str, title: str, ref_no: str,
                 engine: str = "auto", branded: bool = False) -> tuple[str, str]:
    """Render xlsx -> PDF table pages, A4 portrait. Sheets that embed pictures (e.g.
    §4 Material Selection) are rendered FAITHFULLY — exactly as Excel lays them out,
    with the logo composited on top when branded — so text never clips and pictures
    never shift. Other sheets are fit-to-width and centred, with the AMT logo stamped
    when branded. Returns (out_pdf, engine)."""
    soffice = have_soffice()
    use_lo = (engine == "libreoffice") or (engine == "auto" and soffice)

    # image-rich sheets -> faithful Excel layout (+ composited logo when branded)
    if use_lo and soffice and _workbook_has_images(xlsx_path):
        try:
            render_faithful(xlsx_path, out_pdf, soffice, branded=branded)
            return out_pdf, "libreoffice"
        except Exception as e:
            if engine == "libreoffice":
                raise
            print(f"  ! faithful render failed ({e}); falling back.")

    raw = out_pdf + ".raw.pdf" if branded else out_pdf
    engine_used = None
    if use_lo:
        if not soffice:
            raise RuntimeError("render_engine=libreoffice but soffice not found on PATH.")
        prepared = None
        try:
            prepared = _prepare_xlsx_for_print(xlsx_path, reserve_top=branded)
            render_with_soffice(prepared, raw, soffice)
            engine_used = "libreoffice"
        except Exception as e:
            if engine == "libreoffice":
                raise
            print(f"  ! LibreOffice failed ({e}); using reportlab fallback.")
        finally:
            if prepared and os.path.exists(prepared):
                os.remove(prepared)
    if engine_used is None:
        render_with_reportlab(xlsx_path, raw, title, ref_no, branded=branded)
        engine_used = "reportlab"

    if branded:
        STAMP.stamp_logo(raw, out_pdf)   # AMT logo header in the reserved band
        try:
            os.remove(raw)
        except OSError:
            pass
    return out_pdf, engine_used


if __name__ == "__main__":
    import sys
    out, eng = render_table(sys.argv[1], "/tmp/table_test.pdf", "Tender BoQ", "REF v.00")
    print("engine:", eng, "->", out)

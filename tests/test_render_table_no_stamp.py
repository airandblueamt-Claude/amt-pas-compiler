import os, sys, inspect, tempfile
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))
import openpyxl, render_tables as RT

assert "brand" not in inspect.signature(RT.render_table).parameters, "drop brand param"

fd, src = tempfile.mkstemp(suffix=".xlsx"); os.close(fd)
wb = openpyxl.Workbook(); ws = wb.active
ws["A1"] = "Item"; ws["B1"] = "Desc"; ws["A2"] = 1; ws["B2"] = "Ceiling speaker 6W"
wb.save(src)
out = tempfile.mkstemp(suffix=".pdf")[1]
res, eng = RT.render_table(src, out, "BOQ", "REF-001 v.00")
import fitz
d = fitz.open(res)
assert d.page_count >= 1
assert "Ceiling speaker 6W" in d[0].get_text()
print("OK", eng)

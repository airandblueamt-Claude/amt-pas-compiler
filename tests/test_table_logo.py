"""
A BOQ table renders faithfully (content untouched). The AMT logo header is stamped
ONLY when branded=True (AMT-authored sections §2/3/4); unbranded sections stay clean.
Run: python3 tests/test_table_logo.py
"""
import os, sys, inspect, tempfile
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))
import openpyxl, render_tables as RT
import fitz

# render_table is branding-aware via `branded` (not the old `brand`)
params = inspect.signature(RT.render_table).parameters
assert "brand" not in params, "no legacy brand param"
assert "branded" in params, "render_table must accept branded"

fd, src = tempfile.mkstemp(suffix=".xlsx"); os.close(fd)
wb = openpyxl.Workbook(); ws = wb.active
ws["A1"] = "Item"; ws["B1"] = "Desc"; ws["A2"] = 1; ws["B2"] = "Ceiling speaker 6W"
wb.save(src)

# --- branded section: logo stamped in the reserved top band ---
out_b = tempfile.mkstemp(suffix=".pdf")[1]
RT.render_table(src, out_b, "BOQ", "REF-001 v.00", branded=True)
db = fitz.open(out_b)
assert "Ceiling speaker 6W" in db[0].get_text(), "table text must survive (branded)"
assert len(db[0].get_images()) >= 1, "logo header should be stamped when branded"
top_band = RT.A.logo_header_reserve() + 8
in_band = [r for img in db[0].get_images(full=True)
           for r in db[0].get_image_rects(img[0]) if r.y0 < top_band]
assert in_band, "stamped logo should sit in the reserved top margin"

# --- unbranded section: faithful, NO logo ---
out_u = tempfile.mkstemp(suffix=".pdf")[1]
RT.render_table(src, out_u, "BOQ", "REF-001 v.00", branded=False)
du = fitz.open(out_u)
assert "Ceiling speaker 6W" in du[0].get_text(), "table text must survive (unbranded)"
assert len(du[0].get_images()) == 0, "no logo should be stamped when unbranded"

print("OK")

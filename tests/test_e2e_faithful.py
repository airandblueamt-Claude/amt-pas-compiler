import os, sys, tempfile
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))
import openpyxl, fitz
import discover as DISC, assemble as ASM, pas_spec as SPEC, build_pas as BUILD
from openpyxl.styles import Font

work = tempfile.mkdtemp(); inp = os.path.join(work, "in")
secs = ["1-BOQ","2-Mat","3-Trace","4-Sel","5-Cat","8-SLD"]
for d in secs: os.makedirs(os.path.join(inp, d))

def make_sheet(path):
    wb=openpyxl.Workbook(); ws=wb.active
    ws.append(["Item","Tender Description","Qty","Brand","Part No.","Selected Product Description"])
    ws.append([4,"وحدة محطة نداء مكتبية كاملة بجميع التوصيلات","55","TOA","BM-2000","Paging Station 19 Buttons with microphone, desktop mounted"])
    ws.append([None,None,None,"TOA","BM-210P","10 Button additional keys Expansion unit"])
    ws.merge_cells("A2:A3"); ws.merge_cells("B2:B3"); ws.merge_cells("C2:C3")
    ws["B2"].font = Font(name="Times New Roman", size=12)
    for col,w in zip("ABCDEF",[5,38,5,8,12,40]):
        ws.column_dimensions[col].width=w
    wb.save(path)
make_sheet(os.path.join(inp,"1-BOQ/Tender BOQ.xlsx"))
make_sheet(os.path.join(inp,"2-Mat/Mat.xlsx"))
make_sheet(os.path.join(inp,"3-Trace/Trace.xlsx"))
make_sheet(os.path.join(inp,"4-Sel/Sel.xlsx"))
for p,t in [("5-Cat/ds.pdf","TOA Datasheet"),("8-SLD/sld.pdf","Single Line Diagram")]:
    d=fitz.open(); d.new_page(width=595,height=842).insert_text((60,80),t); d.save(os.path.join(inp,p)); d.close()

cfg = {"ref_no":"E2E","version":"00","date":"16-June-2026",
       "project_title_en":"Test","project_title_ar":"اختبار",
       "signoff":{"prepared_by":{"role_en":"x","initials":"A"},
                  "checked_by":{"role_en":"x","initials":"B"},
                  "approved_by":{"role_en":"x","initials":"C"}},
       "revision":{"author":"A","remarks":"r"},
       "input_dir":inp, "output_pdf":os.path.join(work,"out.pdf"),
       "missing_section_mode":"placeholder"}

assert BUILD.validate_config(cfg) == []
tpl = SPEC.resolve_template(cfg)
manifest = DISC.discover(inp, tpl["sections"])
assert not manifest["errors"], manifest["errors"]
qa = ASM.build(manifest, cfg, tpl)
assert qa["consistent"], qa

d = fitz.open(cfg["output_pdf"])
sizes = {(round(p.rect.width), round(p.rect.height)) for p in d}
assert (595,842) in sizes, sizes
alltext = "".join(p.get_text() for p in d)
assert "Paging Station 19 Buttons with microphone" in alltext
# RTL text extracts from the soffice-rendered PDF layer with reordered glyphs/joins,
# so the full contiguous phrase never matches even though it renders correctly.
# Assert the distinctive Arabic words all survived instead.
for w in ["محطة","نداء","مكتبية","كاملة","بجميع"]:
    assert w in alltext, (w, alltext)
print("pages:", d.page_count, "sizes:", sizes)
print("OK")

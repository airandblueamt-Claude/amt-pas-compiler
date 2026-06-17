"""
_prepare_xlsx_for_print rules:
  * FONTS and cell VALUES are never changed (faithful) on any sheet.
  * A sheet WITH images: each picture is PINNED to its display size (a fixed-size
    OneCellAnchor) so later height changes can't squash it; wrapping is turned on and
    image rows GROW (never shrink below the picture) so long text can't clip.
  * An image-FREE sheet: wrapping is turned on and manual row heights are cleared so
    LibreOffice auto-fits each row exactly.
  * page setup is always normalized to A4 fit-to-width.
Run: python3 tests/test_prepare_faithful.py
"""
import os, sys, tempfile
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as XLImage
import render_tables as RT

LOGO = os.path.join(os.path.dirname(__file__), "..", "assets", "amt-logo.png")

# --- sheet WITH an image -> font preserved, picture pinned, row grows for text ---
fd, src = tempfile.mkstemp(suffix=".xlsx"); os.close(fd)
wb = openpyxl.Workbook(); ws = wb.active
ws["A1"] = "Item"
ws["B2"] = "A long Arabic-style description that must wrap and not clip in its row"
ws["B2"].font = Font(name="Times New Roman", size=13, bold=True)
ws["B2"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
ws.row_dimensions[2].height = 22.0
ws.add_image(XLImage(LOGO), "D2")
wb.save(src)

p = openpyxl.load_workbook(RT._prepare_xlsx_for_print(src)).active
assert p["B2"].font.name == "Times New Roman", p["B2"].font.name   # font never changed
assert p["B2"].font.size == 13, p["B2"].font.size
assert p["B2"].font.bold is True, "font weight never changed"
assert p["B2"].alignment.wrap_text is True, "image sheet: wrap turned on so text can't clip"
# the image row grew to fit its text/picture and was never left too short to clip
assert p.row_dimensions[2].height is not None, "image row keeps an explicit height"
assert p.row_dimensions[2].height >= 22.0, "image row never shrinks below original"
# the picture is pinned to a fixed display size (OneCellAnchor with a positive extent)
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor
anch = p._images[0].anchor
assert isinstance(anch, OneCellAnchor), f"picture pinned to fixed-size anchor, got {type(anch).__name__}"
assert anch.ext.cx > 0 and anch.ext.cy > 0, "pinned picture keeps a positive size (not squashed)"
assert p.page_setup.fitToWidth == 1

# --- image-FREE sheet -> wrap on, row heights auto-fit, font still preserved ---
fd, src2 = tempfile.mkstemp(suffix=".xlsx"); os.close(fd)
wb2 = openpyxl.Workbook(); ws2 = wb2.active
ws2["A1"] = "Item"; ws2["B2"] = "A long description in a too-short fixed row"
ws2["B2"].font = Font(name="Times New Roman", size=13)
ws2["B2"].alignment = Alignment(wrap_text=False)
ws2.row_dimensions[2].height = 12.0
wb2.save(src2)

q = openpyxl.load_workbook(RT._prepare_xlsx_for_print(src2)).active
assert q["B2"].font.name == "Times New Roman", "font never changed even when auto-fitting"
assert q["B2"].alignment.wrap_text is True, "image-free sheet: wrap turned on"
assert q.row_dimensions[2].height is None, "image-free sheet: manual height cleared (auto-fit)"

print("OK")
